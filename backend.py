from flask import Flask, request, jsonify
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)
xlsx_filename = "keys.xlsx"

def check_existing_key(api_key):
    try:
        wb = openpyxl.load_workbook(xlsx_filename)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
            stored_api_key, = row
            if stored_api_key == api_key:
                return True
        return False
    except FileNotFoundError:
        return False
def check_and_update_machine_code(api_key, machine_code):
    try:
        wb = openpyxl.load_workbook(xlsx_filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["API Key", "Machine Code", "Activation Time", "Additional Data", "Status"])

    sheet = wb.active
    activation_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
        stored_api_key, stored_machine_code, _, _, _, status = row
        if stored_api_key.value == api_key:
            if status.value == "{false}":
                stored_machine_code.value = machine_code
                status.value = "{true}"
                row[2].value = activation_time
                wb.save(xlsx_filename)
                return jsonify({"message": "机器码已更新"}), 402
            elif status.value == "{true}" and stored_machine_code.value == machine_code:
                return jsonify({"message": "机器码验证通过"}), 200
    if not check_existing_key(api_key):
        sheet.append([api_key, machine_code, activation_time, None, "{true}"])
        wb.save(xlsx_filename)
        return jsonify({"message": "首次使用"}), 402
    return jsonify({"message": "无效的机器码"}), 401

@app.route('/login', methods=['POST'])
def login():
    api_key = request.headers.get('Authorization').replace('Bearer ', '')
    machine_code = request.form.get('machine_code')

    if not api_key:
        return jsonify({"message": "Invalid request"}), 400

    if not check_existing_key(api_key):
        return jsonify({"message": "Invalid API Key"}), 401

    return check_and_update_machine_code(api_key, machine_code)



if __name__ == '__main__':
    if not os.path.isfile(xlsx_filename):
        openpyxl.Workbook().save(xlsx_filename)
#公网请使用默认 0.0.0.0 内网则需要输入对应的IP地址
    app.run(host='0.0.0.0', port=50000, debug=True)
